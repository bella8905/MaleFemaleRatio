#!/usr/bin/env python

# import openpyxl for excel file processing
# there are alternatives as pandas...
import openpyxl
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule


# import os for cross platform file path
# windows uses '\', but this makes MacOS angry
import os


def get_number_from_str( _str ):
    if _str.isdigit() and int(_str) > 0:
        return int(_str)
    else:
        return 0


def fix_ratio(_ratio):
    valid = False
    male = 0
    female = 0

    # strip whitespece
    ratio = _ratio.replace(' ', '')

    delimiter = ':'
    # (a/b) or (a:b)
    ratio = ratio.replace('/', delimiter).replace('+', delimiter).replace('--', delimiter).replace('-', delimiter).\
        replace('_', delimiter).replace('¦', delimiter).replace('\\', delimiter).replace('%', delimiter).replace(':', delimiter).\
        replace('：', delimiter).replace('／', delimiter).replace('╱', delimiter).replace('；', delimiter).replace('－', delimiter).\
        replace('比', delimiter)
    ratio = ratio.split(delimiter)
    if len(ratio) == 2:
        male = get_number_from_str( ratio[0] )
        if male > 0:
            female = get_number_from_str( ratio[1] )
            if female > 0:
                valid = True

    return valid, male, female


# wb = openpyxl.load_workbook( filename="data/MaleFemaleRatio.xlsx", data_only=True )
cwd = os.getcwd()
# print( "current directory:" + cwd )
in_file = os.path.join(cwd, "data", "MaleFemaleRatio.xlsx")
out_file = os.path.join(cwd, "out", "MaleFemaleRatioUpdated.xlsx")

# read from xlsx
# data-only: ignore formulas
wb = openpyxl.load_workbook(filename=in_file, data_only=True)
# wb_sheets = wb.sheetnames
# print(wb_sheets)
ws = wb["需要计算的男女比"]

# print( "1".isdigit() )
# print( "-1".isdigit() )
# print( "1.2".isdigit() )
# print( "  1".replace(' ','').isdigit() )
# print( " 1 4 ".replace(' ','').isdigit() )

# read from each row:
# skip the first row, it the titles
# only read from the first col
ws["C1"] = "male"
ws["D1"] = "female"
ws["E1"] = "ratio"
print( len(ws['A']) )

for row in ws.iter_rows(min_row=2):
    ratio = row[0].value
    if ratio is None:
        ratio = ""
    valid, male, female = fix_ratio(ratio)
    if not valid:
        print("not valid: " + ratio)
    row[2].value = male if valid else 0
    row[3].value = female if valid else 0
    row[4].value = male / female if valid else -1
    row[4].value = float("{0:.4f}".format(row[4].value))

# highlight the invalid data
red_background = PatternFill(bgColor=colors.RED)
diff_style = DifferentialStyle(fill=red_background)
rule = Rule(type="expression", dxf=diff_style)
rule.formula = ["$E1<0"]
ws.conditional_formatting.add("A1:E{0}".format(len(ws['A'])), rule)

wb.save(filename=out_file)

print(float("1.2a"))
